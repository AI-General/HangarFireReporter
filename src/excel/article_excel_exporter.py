import os
import pandas as pd
from src.llm.language import translate_text
from src.db import get_articles
from src.config import Config
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

class ArticleExcelExporter:
    HEADERS = [
        "Date of Incident",
        "Airport / Hangar Name",
        "Country / Region",
        "Brief Summary",
        "Source Link(s)",
        "Language",
        "Origin Title"
    ]
    MAX_COL_WIDTH = 50  # Maximum column width

    def __init__(self):
        self.output_path = Config.REPORT_FILE_PATH or "reports/hangar_fire_report.xlsx"

    def export_articles_to_excel(self):
        # Fetch articles from Supabase
        articles = get_articles()
        if not articles:
            print(f"No articles found.")
            return

        # Prepare new data
        new_rows = []
        for article in articles:
            # Format URLs as comma-separated string (no brackets)
            urls = article.get("url", [])
            if len(urls) > 3:
                urls = urls[:3]
            if isinstance(urls, list):
                url_str = ", ".join(urls)
            else:
                url_str = str(urls)
            
            language = article.get("language", "en")
            summary = article.get("description") or article.get("title") 
            
            if language != "en":
                summary = translate_text(summary, "en", language)   
            
            row = {
                "Date of Incident": article.get("publishedAt", ""),
                "Airport / Hangar Name": article.get("airport_hangar_name", ""),
                "Country / Region": article.get("location", ""),
                "Brief Summary": summary,
                "Source Link(s)": url_str,
                "Language": language,
                "Origin Title": article.get("title", "") if language != "en" else "",
            }
            new_rows.append(row)
        new_df = pd.DataFrame(new_rows, columns=self.HEADERS)

        # If file exists, update it
        if os.path.exists(self.output_path):
            old_df = pd.read_excel(self.output_path)
            # Remove duplicates based on Source Link(s)
            combined_df = pd.concat([new_df, old_df], ignore_index=True)
            combined_df.drop_duplicates(subset=["Source Link(s)"], keep="last", inplace=True)
            combined_df.reset_index(drop=True, inplace=True)
        else:
            combined_df = new_df

        # Sort by Date of Incident (descending, most recent first)
        combined_df.sort_values(by="Date of Incident", ascending=False, inplace=True)

        # Write to Excel
        combined_df.to_excel(self.output_path, index=False)

        # Post-process with openpyxl for styling
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Style header row
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        header_font = Font(bold=True)
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # Adjust column widths
        for col_idx, header in enumerate(self.HEADERS, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            max_length = min(max_length + 2, self.MAX_COL_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length

        # Format URLs as clickable hyperlinks and color them blue
        url_col_idx = self.HEADERS.index("Source Link(s)") + 1
        blue_font = Font(color="0000EE", underline="single")
        for row in ws.iter_rows(min_row=2, min_col=url_col_idx, max_col=url_col_idx):
            for cell in row:
                if cell.value:
                    urls = [u.strip() for u in str(cell.value).split(",") if u.strip()]
                    if urls:
                        # If multiple URLs, join with comma and space, each as hyperlink
                        display = []
                        for u in urls:
                            cell.value = u  # Set to first URL for hyperlink
                            cell.hyperlink = u
                            cell.font = blue_font
                            display.append(u)
                        if len(display) > 1:
                            # If multiple, show as comma-separated, but only first is clickable
                            cell.value = ", ".join(display)
                            # openpyxl only supports one hyperlink per cell
        wb.save(self.output_path)
        print(f"Exported {len(new_df)} new/updated articles to {self.output_path}") 